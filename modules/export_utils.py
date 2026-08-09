import pandas as pd
import io
import re
import html
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from modules.data_processing import formata_campo_html

def identificar_icone_folium(row, colunas):
    tipo_str = str(row.get('TIPO LIGACAO', '')) + str(row.get('SERVICO', '')) + str(row.get('TIPO NOTA', ''))
    tipo_str = tipo_str.upper()
    if row.get('PROTOCOLO') == 'RETORNO_BASE': return 'home'
    if row.get('PROTOCOLO') == 'PAUSA_ALMOCO': return 'cutlery'
    if 'NOVA' in tipo_str or 'LIGACAO' in tipo_str or 'UNI' in tipo_str or 'UNR' in tipo_str: return 'bolt'
    if 'MANUT' in tipo_str or 'REPARO' in tipo_str: return 'wrench'
    if 'INSP' in tipo_str or 'VISTORIA' in tipo_str: return 'eye-open'
    return 'info-sign'

def renderizar_painel_lateral(cap_eq, obras_prontas, eq_sel, cap_tot):
    return f'''
    <label style="font-size: 13.5px; font-weight: 700; color: #0D256C; margin-bottom: 4px; display: block; margin-top: 10px;">Capacidade da Rota (Por Equipe):</label>
    <div style="background-color: #f8f9fa; border: 1px solid #e2e8f0; border-radius: 6px; padding: 8px 12px; margin-bottom: 12px; color: #d9534f; font-weight: 800; font-size: 15px; text-align: center;">
        {cap_eq} Obras
    </div>
    
    <label style="font-size: 13.5px; font-weight: 700; color: #0D256C; margin-bottom: 4px; display: block;">Obras Prontas p/ Roteirizar:</label>
    <div style="background-color: #edf7ed; border: 1px solid #c8e6c9; border-radius: 6px; padding: 8px 12px; margin-bottom: 12px; color: #2e7d32; font-weight: 800; font-size: 15px; text-align: center;">
        {obras_prontas} Obras
    </div>
    
    <label style="font-size: 13.5px; font-weight: 700; color: #0D256C; margin-bottom: 4px; display: block;">Equipes Selecionadas:</label>
    <div style="background-color: #f8f9fa; border: 1px solid #e2e8f0; border-radius: 6px; padding: 8px 12px; margin-bottom: 12px; color: #d9534f; font-weight: 800; font-size: 15px; text-align: center;">
        {eq_sel} Equipes
    </div>
    
    <label style="font-size: 13.5px; font-weight: 700; color: #0D256C; margin-bottom: 4px; display: block;">Quantidade Total Projetada:</label>
    <div style="background-color: #f8f9fa; border: 1px solid #e2e8f0; border-radius: 6px; padding: 8px 12px; margin-bottom: 5px; color: #d9534f; font-weight: 800; font-size: 15px; text-align: center;">
        {cap_tot} Obras Max.
    </div>
    '''

def gerar_csv_autocad_proj(df_routed):
    """Gera um arquivo CSV limpo no padrão Topografia (Ponto, X, Y, Z, Descrição) para importar no Proj+/AutoCAD"""
    df_cad = pd.DataFrame()
    df_real = df_routed[~df_routed['PROTOCOLO'].isin(['RETORNO_BASE', 'PAUSA_ALMOCO'])].copy()
    
    df_cad['PONTO'] = df_real['PROTOCOLO']
    df_cad['ESTE_LONGITUDE'] = df_real['LONGITUDE']
    df_cad['NORTE_LATITUDE'] = df_real['LATITUDE']
    df_cad['ELEVACAO_Z'] = 0
    
    # Tratamento seguro caso a coluna de postes previstos não exista
    if 'POSTES PREVISTOS' in df_real.columns:
        postes_str = " | Postes Prev: " + df_real['POSTES PREVISTOS'].fillna(0).astype(int).astype(str)
    else:
        postes_str = ""
        
    df_cad['DESCRICAO'] = df_real['NOME'] + postes_str + " | Eq: " + df_real['BASE_ATRIBUIDA']
    
    return df_cad.to_csv(index=False, sep=';').encode('utf-8-sig')

def gerar_excel_bytes(df, col_prioridade, colunas_originais=None):
    df_export = df.copy()
    if 'PROTOCOLO' in df_export.columns: 
        df_export = df_export[~df_export['PROTOCOLO'].isin(['RETORNO_BASE', 'PAUSA_ALMOCO'])]
        
    unpacked_rows = []
    for _, row in df_export.iterrows():
        if '_ORIGINAL_ROWS' in row and isinstance(row['_ORIGINAL_ROWS'], list):
            for orig in row['_ORIGINAL_ROWS']:
                new_row = orig.copy()
                link_maps = f"https://www.google.com/maps/search/?api=1&query={new_row.get('LATITUDE','')},{new_row.get('LONGITUDE','')}"
                new_row['LINK_NAVEGACAO_OFFLINE'] = link_maps
                
                for vrp_col in ['NOME_DIA', 'ORDEM', 'SEMANA', 'DIA', 'PERIODO', 'DISTANCIA_PONTO_ANTERIOR_KM', 'ALERTA_TOPOLOGIA', 'DISTANCIA_PROXIMO_PONTO_KM', 'TEMPO_VIAGEM_MINUTOS', 'HORA_INICIO', 'HORA_FIM', 'SUPER_PONTO', 'BASE_ATRIBUIDA', 'PRIORIDADE', 'DISTANCIA_REDE_METROS', 'POSTES PREVISTOS', 'LATITUDE_REDE', 'LONGITUDE_REDE']:
                    if vrp_col in row:
                        new_row[vrp_col] = row[vrp_col]
                unpacked_rows.append(new_row)
        else:
            row_dict = row.to_dict()
            row_dict['LINK_NAVEGACAO_OFFLINE'] = f"https://www.google.com/maps/search/?api=1&query={row_dict.get('LATITUDE','')},{row_dict.get('LONGITUDE','')}"
            unpacked_rows.append(row_dict)
            
    df_export = pd.DataFrame(unpacked_rows)
    for col in ['ROTA_GEOMETRIA', 'STATUS LIST', 'INICIO AVARIA', 'STATUS ATUAL (LEVANTAMENTO)', 'DESCRICAO', '_HORA_INICIO_DT', '_HORA_FIM_DT', '_ORIGINAL_ROWS', '_ORIGEM_BASE', 'MUN_LIMPO_CALC']:
        if col in df_export.columns: df_export = df_export.drop(columns=[col], errors='ignore')

    if colunas_originais:
        cols_atuais = df_export.columns.tolist()
        cols_originais_validas = [c for c in colunas_originais if c in cols_atuais]
        cols_novas_geradas = [c for c in cols_atuais if c not in colunas_originais]
        df_export = df_export[cols_originais_validas + cols_novas_geradas]
        
    buf_xl = io.BytesIO()
    with pd.ExcelWriter(buf_xl, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Roteiro')
        ws = writer.sheets['Roteiro']
        header_fill = PatternFill(start_color='0D256C', end_color='0D256C', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        col_types = {}
        for col_idx, col_name in enumerate(df_export.columns, 1):
            col_letter = get_column_letter(col_idx)
            col_name_upper = str(col_name).upper()
            if any(x in col_name_upper for x in ['NOME', 'CLIENTE', 'ENDEREÇO', 'INFORMAÇ']): ws.column_dimensions[col_letter].width = 45.0
            elif any(x in col_name_upper for x in ['PROTOCOLO', 'MUNICIPIO', 'BASE', 'LOCALIDADE', 'LINK_NAVEGACAO_OFFLINE']): ws.column_dimensions[col_letter].width = 25.0
            else: ws.column_dimensions[col_letter].width = 18.0
                
            if col_name_upper in ['NOME_DIA', 'ORDEM', 'SEMANA', 'DIA', 'PERIODO', 'DISTANCIA_PONTO_ANTERIOR_KM', 'DISTANCIA_PROXIMO_PONTO_KM', 'TEMPO_VIAGEM_MINUTOS', 'PRIORIDADE', 'HORA_INICIO', 'HORA_FIM', 'DISTANCIA_REDE_METROS', 'POSTES PREVISTOS', 'LATITUDE_REDE', 'LONGITUDE_REDE']:
                col_types[col_idx] = center_align
            else: col_types[col_idx] = left_align

        red_font = Font(color="FF0000", bold=True)
        prio_idx = df_export.columns.get_loc('PRIORIDADE') + 1 if 'PRIORIDADE' in df_export.columns else None
        prio_target_idx = df_export.columns.get_loc(col_prioridade) + 1 if col_prioridade in df_export.columns else None

        for row_idx in range(2, len(df_export) + 2):
            for col_idx in range(1, len(df_export.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = col_types.get(col_idx, left_align)
                if df_export.columns[col_idx - 1] == 'LINK_NAVEGACAO_OFFLINE' and cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")
            if prio_idx and ws.cell(row=row_idx, column=prio_idx).value == "Sim":
                ws.cell(row=row_idx, column=prio_idx).font = red_font
                if prio_target_idx and col_prioridade != "Nenhuma":
                    try: ws.cell(row=row_idx, column=prio_target_idx).font = red_font
                    except: pass
                    
        if 'SUPER_PONTO' in df_export.columns:
            try:
                idx_super = df_export.columns.get_loc('SUPER_PONTO') + 1
                yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                for row_idx in range(2, len(df_export) + 2):
                    val = str(ws.cell(row=row_idx, column=idx_super).value)
                    if val.startswith("SIM"):
                        for col_idx_f in range(1, len(df_export.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx_f).fill = yellow_fill
            except: pass
            
        if 'ALERTA_TOPOLOGIA' in df_export.columns:
            try:
                idx_alerta = df_export.columns.get_loc('ALERTA_TOPOLOGIA') + 1
                orange_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
                for row_idx in range(2, len(df_export) + 2):
                    val = str(ws.cell(row=row_idx, column=idx_alerta).value)
                    if "Suspeita" in val:
                        ws.cell(row=row_idx, column=idx_alerta).fill = orange_fill
                        ws.cell(row=row_idx, column=idx_alerta).font = Font(color="8B0000", bold=True)
            except: pass
                    
    return buf_xl.getvalue()

def gerar_excel_resumo_bytes(df):
    buf_xl = io.BytesIO()
    with pd.ExcelWriter(buf_xl, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Resumo')
        ws = writer.sheets['Resumo']
        header_fill = PatternFill(start_color='0D256C', end_color='0D256C', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            if str(col_name).upper() == 'LEVANTADOR':
                ws.column_dimensions[col_letter].width = 45.0
                col_align = left_align
            else:
                ws.column_dimensions[col_letter].width = 22.0
                col_align = center_align
            for row_idx in range(2, len(df) + 2): ws.cell(row=row_idx, column=col_idx).alignment = col_align
    return buf_xl.getvalue()

def gerar_kml_agrupado(df_rota, bases_records, doc_name, cols_exibir, lista_todas_bases=None, tipo_periodo="Dia"):
    if lista_todas_bases is None: lista_todas_bases = df_rota['BASE_ATRIBUIDA'].unique().tolist()
        
    kml_lines = [f'''<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>{doc_name}</name>
  <Style id="linha-rota-contorno"><LineStyle><color>ff000000</color><width>8</width></LineStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>
  <Style id="linha-ligacao-rede"><LineStyle><color>8800ffff</color><width>2</width></LineStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>

  <Style id="style-blue-n"><IconStyle><scale>1.1</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>
  <Style id="style-blue-h"><IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle><LabelStyle><scale>1.0</scale><color>ffffffff</color></LabelStyle></Style>
  <StyleMap id="icon-blue"><Pair><key>normal</key><styleUrl>#style-blue-n</styleUrl></Pair><Pair><key>highlight</key><styleUrl>#style-blue-h</styleUrl></Pair></StyleMap>

  <Style id="style-red-n"><IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>
  <Style id="style-red-h"><IconStyle><scale>1.5</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-blank.png</href></Icon><hotSpot x="32" xunits="pixels" y="64" yunits="insetPixels"/></IconStyle><LabelStyle><scale>1.0</scale><color>ffffffff</color></LabelStyle></Style>
  <StyleMap id="icon-red"><Pair><key>normal</key><styleUrl>#style-red-n</styleUrl></Pair><Pair><key>highlight</key><styleUrl>#style-red-h</styleUrl></Pair></StyleMap>

  <Style id="style-green-n"><IconStyle><scale>1.2</scale><Icon><href>https://maps.google.com/mapfiles/kml/shapes/homegardenbusiness.png</href></Icon></IconStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>
  <Style id="style-green-h"><IconStyle><scale>1.4</scale><Icon><href>https://maps.google.com/mapfiles/kml/shapes/homegardenbusiness.png</href></Icon></IconStyle><LabelStyle><scale>1.0</scale><color>ffffffff</color></LabelStyle></Style>
  <StyleMap id="icon-green"><Pair><key>normal</key><styleUrl>#style-green-n</styleUrl></Pair><Pair><key>highlight</key><styleUrl>#style-green-h</styleUrl></Pair></StyleMap>

  <Style id="style-yellow-n"><IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/ylw-blank.png</href></Icon></IconStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>
  <Style id="style-yellow-h"><IconStyle><scale>1.5</scale><Icon><href>http://maps.google.com/mapfiles/kml/paddle/ylw-blank.png</href></Icon></IconStyle><LabelStyle><scale>1.0</scale><color>ffffffff</color></LabelStyle></Style>
  <StyleMap id="icon-yellow"><Pair><key>normal</key><styleUrl>#style-yellow-n</styleUrl></Pair><Pair><key>highlight</key><styleUrl>#style-yellow-h</styleUrl></Pair></StyleMap>''']

    kml_cores = ['ff4b19e6', 'ffd4bc00', 'ffb5513f', 'ff889600', 'ff0098ff', 'ffb0279c', 'ff39dccd', 'ff631ee9', 'ff3bebff', 'ff485579']
    for idx, b_nome in enumerate(lista_todas_bases):
        cor_kml = kml_cores[idx % len(kml_cores)]
        nome_limpo = re.sub(r'[^A-Za-z0-9_]', '', str(b_nome))
        kml_lines.append(f'  <Style id="rota-centro-{nome_limpo}"><LineStyle><color>{cor_kml}</color><width>5</width></LineStyle><LabelStyle><scale>0</scale><color>00ffffff</color></LabelStyle></Style>')

    for base_nome in df_rota['BASE_ATRIBUIDA'].unique():
        df_base = df_rota[df_rota['BASE_ATRIBUIDA'] == base_nome]
        base_ref = next((b for b in bases_records if b['LEVANTADOR'] == base_nome), None)
        b_lat, b_lon = float(str(base_ref['LATITUDE']).replace(',','.')), float(str(base_ref['LONGITUDE']).replace(',','.'))
        res_nome = str(base_ref.get('RESIDENCIA', base_nome))
        nome_limpo_base = re.sub(r'[^A-Za-z0-9_]', '', str(base_nome))

        kml_lines.append(f'  <Folder>\n    <name>Levantador: {html.escape(str(base_nome))}</name>')
        kml_lines.append(f'    <Placemark><name>BASE: {html.escape(str(res_nome))}</name><styleUrl>#icon-green</styleUrl><Point><coordinates>{b_lon},{b_lat},0</coordinates></Point></Placemark>')

        if tipo_periodo == "Semana":
            for semana in df_base['SEMANA'].unique():
                df_semana = df_base[df_base['SEMANA'] == semana]
                kml_lines.append(f'    <Folder>\n      <name>Semana {semana}</name>')
                for dia in df_semana['DIA'].unique():
                    df_dia = df_semana[df_semana['DIA'] == dia].copy().sort_values(by='ORDEM')
                    nome_dia_kml = df_dia.iloc[0].get('NOME_DIA', f"Dia {dia}") if not df_dia.empty else f"Dia {dia}"
                    kml_lines.append(f'      <Folder>\n        <name>{nome_dia_kml}</name>')
                    coords_linha_kml = []
                    
                    for r in df_dia.to_dict('records'):
                        lon, lat = str(r.get('LONGITUDE')).replace(',','.'), str(r.get('LATITUDE')).replace(',','.')
                        
                        geometria = r.get('ROTA_GEOMETRIA')
                        if isinstance(geometria, list):
                            coords_linha_kml.extend([f"          {pt_lon},{pt_lat},0" for pt_lon, pt_lat in geometria])
                        else:
                            coords_linha_kml.append(f"          {lon},{lat},0")

                        if r.get('PROTOCOLO') in ['RETORNO_BASE', 'PAUSA_ALMOCO']:
                            continue

                        is_super = str(r.get('SUPER_PONTO', '')).startswith('SIM')
                        
                        if is_super:
                            qtd_str = str(r.get('SUPER_PONTO')).replace('SIM', '').strip()
                            pop_header_bg, pop_header_color = "#FFD700", "#000000"
                            pop_prio_txt = f"🏢 SUPER PONTO {qtd_str}"
                            nome_ponto = f"[PRIORIDADE] [{r.get('ORDEM', 0)}] 🏢 SUPER PONTO {qtd_str}" if r.get('PRIORIDADE') == "Sim" else f"[{r.get('ORDEM', 0)}] 🏢 SUPER PONTO {qtd_str}"
                            style_url = "#icon-yellow"
                        else:
                            pop_header_bg, pop_header_color = ("#d9534f", "#ffffff") if r.get('PRIORIDADE') == "Sim" else ("#0D256C", "#ffffff")
                            pop_prio_txt = "🚨 OBRA PRIORITÁRIA" if r.get('PRIORIDADE') == "Sim" else "📍 Atendimento Padrão"
                            prot_str = str(r.get('PROTOCOLO', 'N/A'))
                            nome_obra = str(r.get('NOME', ''))
                            if nome_obra.lower() == 'nan': nome_obra = ''
                            separador = " - " if nome_obra else ""
                            tag_prio = "[PRIORIDADE] " if r.get('PRIORIDADE') == "Sim" else ""
                            nome_ponto = f"{tag_prio}[{r.get('ORDEM', 0)}] Doc: {html.escape(prot_str)}{separador}{html.escape(nome_obra)}"
                            style_url = "#icon-red" if r.get('PRIORIDADE') == "Sim" else "#icon-blue"
                        
                        dist_prox = r.get('DISTANCIA_PROXIMO_PONTO_KM', 0.0)
                        dist_rede = r.get('DISTANCIA_REDE_METROS')
                        postes_prev = r.get('POSTES PREVISTOS')
                        rede_lat = r.get('LATITUDE_REDE')
                        rede_lon = r.get('LONGITUDE_REDE')
                        
                        extra_rows_list = []
                        for c in cols_exibir:
                            if c.upper() not in ['PROTOCOLO', 'NOME_DIA', 'SEMANA']:
                                val_html = formata_campo_html(r.get(c, ''))
                                extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555; vertical-align:top; width:35%;'>{html.escape(str(c))}:</td><td style='padding:3px 6px; color:#333;'>{val_html}</td></tr>")
                                
                        if pd.notna(dist_rede):
                            extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555;'>Rede Mais Próxima:</td><td style='padding:3px 6px; color:#17a2b8; font-weight:bold;'>{dist_rede:.1f} Metros</td></tr>")
                        if pd.notna(rede_lat) and pd.notna(rede_lon):
                            extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555;'>Coord. da Rede:</td><td style='padding:3px 6px; color:#e83e8c; font-weight:bold;'>{rede_lat:.6f}, {rede_lon:.6f}</td></tr>")
                        if pd.notna(postes_prev):
                            extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555;'>Postes Previstos:</td><td style='padding:3px 6px; color:#e67e22; font-weight:bold;'>{int(postes_prev)} UN</td></tr>")

                        extra_rows = "".join(extra_rows_list)
                        prot_html = formata_campo_html(r.get('PROTOCOLO', 'N/A'))

                        popup_html = f"""
                        <div style="font-family:sans-serif; width:280px; border-radius:8px; overflow:hidden; box-shadow:0 2px 5px rgba(0,0,0,0.15);">
                            <div style="background:{pop_header_bg}; color:{pop_header_color}; padding:8px 10px; font-size:13px; font-weight:bold;">{pop_prio_txt}</div>
                            <div style="padding:10px; background:#fafafa; font-size:12px;">
                                <table style="width:100%; border-collapse:collapse;">
                                    <tr><td style="padding:3px 6px; font-weight:bold; color:#555; vertical-align:top; width:35%;">Nota/Protocolo:</td><td style="padding:3px 6px; color:#333;">{prot_html}</td></tr>
                                    <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Ordem:</td><td style="padding:3px 6px; color:#333;">{r.get('ORDEM', 0)} ({r.get('NOME_DIA', f'Dia {r.get("DIA", 0)}')})</td></tr>
                                    <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Horário:</td><td style="padding:3px 6px; color:#333;">{r.get('HORA_INICIO', '')} às {r.get('HORA_FIM', '')}</td></tr>
                                    <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Distância Ant.:</td><td style="padding:3px 6px; color:#333;">{r.get('DISTANCIA_PONTO_ANTERIOR_KM', 0)} KM</td></tr>
                                    <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Distância Próx.:</td><td style="padding:3px 6px; color:#333;">{dist_prox} KM</td></tr>
                                    {extra_rows}
                                </table>
                            </div>
                        </div>"""
                        kml_lines.append(f'        <Placemark><name>{nome_ponto}</name><description><![CDATA[{popup_html}]]></description><styleUrl>{style_url}</styleUrl><Point><coordinates>{lon},{lat},0</coordinates></Point></Placemark>')
                        
                        # --- LINHA GUIA PARA A REDE ELÉTRICA MAIS PRÓXIMA ---
                        if pd.notna(rede_lat) and pd.notna(rede_lon):
                            rede_lat_str = str(rede_lat).replace(',', '.')
                            rede_lon_str = str(rede_lon).replace(',', '.')
                            kml_lines.append(f'        <Placemark><name>Guia de Rede: {dist_rede:.1f}m</name><styleUrl>#linha-ligacao-rede</styleUrl><LineString><tessellate>1</tessellate><coordinates>{lon},{lat},0 {rede_lon_str},{rede_lat_str},0</coordinates></LineString></Placemark>')

                    kml_str_coords = "\n".join(coords_linha_kml)
                    if kml_str_coords.strip():
                        kml_lines.append(f'        <Placemark><name>Contorno Rota</name><styleUrl>#linha-rota-contorno</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{kml_str_coords}\n            </coordinates></LineString></Placemark>')
                        kml_lines.append(f'        <Placemark><name>Traçado Rota</name><styleUrl>#rota-centro-{nome_limpo_base}</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{kml_str_coords}\n            </coordinates></LineString></Placemark>\n      </Folder>')
                    else:
                        kml_lines.append('      </Folder>')
                kml_lines.append('    </Folder>')
        else:
            for dia in df_base['DIA'].unique():
                df_dia = df_base[df_base['DIA'] == dia].copy().sort_values(by='ORDEM')
                nome_dia_kml = df_dia.iloc[0].get('NOME_DIA', f"Dia {dia}") if not df_dia.empty else f"Dia {dia}"
                
                kml_lines.append(f'      <Folder>\n        <name>{nome_dia_kml}</name>')
                coords_linha_kml = []
                
                for r in df_dia.to_dict('records'):
                    lon, lat = str(r.get('LONGITUDE')).replace(',','.'), str(r.get('LATITUDE')).replace(',','.')
                    
                    geometria = r.get('ROTA_GEOMETRIA')
                    if isinstance(geometria, list):
                        coords_linha_kml.extend([f"          {pt_lon},{pt_lat},0" for pt_lon, pt_lat in geometria])
                    else:
                        coords_linha_kml.append(f"          {lon},{lat},0")

                    if r.get('PROTOCOLO') in ['RETORNO_BASE', 'PAUSA_ALMOCO']:
                        continue

                    is_super = str(r.get('SUPER_PONTO', '')).startswith('SIM')
                    cor_icone = 'orange' if is_super else ('red' if r.get('PRIORIDADE') == "Sim" else 'blue')
                    
                    if is_super:
                        qtd_str = str(r.get('SUPER_PONTO')).replace('SIM', '').strip()
                        pop_header_bg, pop_header_color = "#FFD700", "#000000"
                        pop_prio_txt = f"🏢 SUPER PONTO {qtd_str}"
                        nome_ponto = f"[PRIORIDADE] [{r.get('ORDEM', 0)}] 🏢 SUPER PONTO {qtd_str}" if r.get('PRIORIDADE') == "Sim" else f"[{r.get('ORDEM', 0)}] 🏢 SUPER PONTO {qtd_str}"
                        style_url = "#icon-yellow"
                    else:
                        pop_header_bg, pop_header_color = ("#d9534f", "#ffffff") if r.get('PRIORIDADE') == "Sim" else ("#0D256C", "#ffffff")
                        pop_prio_txt = "🚨 OBRA PRIORITÁRIA" if r.get('PRIORIDADE') == "Sim" else "📍 Atendimento Padrão"
                        prot_str = str(r.get('PROTOCOLO', 'N/A'))
                        nome_obra = str(r.get('NOME', ''))
                        if nome_obra.lower() == 'nan': nome_obra = ''
                        separador = " - " if nome_obra else ""
                        tag_prio = "[PRIORIDADE] " if r.get('PRIORIDADE') == "Sim" else ""
                        nome_ponto = f"{tag_prio}[{r.get('ORDEM', 0)}] Doc: {html.escape(prot_str)}{separador}{html.escape(nome_obra)}"
                        style_url = "#icon-red" if r.get('PRIORIDADE') == "Sim" else "#icon-blue"
                    
                    dist_prox = r.get('DISTANCIA_PROXIMO_PONTO_KM', 0.0)
                    dist_rede = r.get('DISTANCIA_REDE_METROS')
                    postes_prev = r.get('POSTES PREVISTOS')
                    rede_lat = r.get('LATITUDE_REDE')
                    rede_lon = r.get('LONGITUDE_REDE')
                    
                    extra_rows_list = []
                    for c in cols_exibir:
                        if c.upper() not in ['PROTOCOLO', 'NOME_DIA', 'SEMANA']:
                            val_html = formata_campo_html(r.get(c, ''))
                            extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555; vertical-align:top; width:35%;'>{html.escape(str(c))}:</td><td style='padding:3px 6px; color:#333;'>{val_html}</td></tr>")
                            
                    if pd.notna(dist_rede):
                        extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555;'>Rede Mais Próxima:</td><td style='padding:3px 6px; color:#17a2b8; font-weight:bold;'>{dist_rede:.1f} Metros</td></tr>")
                    if pd.notna(rede_lat) and pd.notna(rede_lon):
                        extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555;'>Coord. da Rede:</td><td style='padding:3px 6px; color:#e83e8c; font-weight:bold;'>{rede_lat:.6f}, {rede_lon:.6f}</td></tr>")
                    if pd.notna(postes_prev):
                        extra_rows_list.append(f"<tr><td style='padding:3px 6px; font-weight:bold; color:#555;'>Postes Previstos:</td><td style='padding:3px 6px; color:#e67e22; font-weight:bold;'>{int(postes_prev)} UN</td></tr>")

                    extra_rows = "".join(extra_rows_list)
                    prot_html = formata_campo_html(r.get('PROTOCOLO', 'N/A'))

                    popup_html = f"""
                    <div style="font-family:sans-serif; width:280px; border-radius:8px; overflow:hidden; box-shadow:0 2px 5px rgba(0,0,0,0.15);">
                        <div style="background:{pop_header_bg}; color:{pop_header_color}; padding:8px 10px; font-size:13px; font-weight:bold;">{pop_prio_txt}</div>
                        <div style="padding:10px; background:#fafafa; font-size:12px;">
                            <table style="width:100%; border-collapse:collapse;">
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555; vertical-align:top; width:35%;">Nota/Protocolo:</td><td style="padding:3px 6px; color:#333;">{prot_html}</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Ordem:</td><td style="padding:3px 6px; color:#333;">{r.get('ORDEM', 0)} ({r.get('NOME_DIA', f'Dia {r.get("DIA", 0)}')})</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Horário:</td><td style="padding:3px 6px; color:#333;">{r.get('HORA_INICIO', '')} às {r.get('HORA_FIM', '')}</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Distância Ant.:</td><td style="padding:3px 6px; color:#333;">{r.get('DISTANCIA_PONTO_ANTERIOR_KM', 0)} KM</td></tr>
                                <tr><td style="padding:3px 6px; font-weight:bold; color:#555;">Distância Próx.:</td><td style="padding:3px 6px; color:#333;">{dist_prox} KM</td></tr>
                                {extra_rows}
                            </table>
                        </div>
                    </div>"""
                    kml_lines.append(f'        <Placemark><name>{nome_ponto}</name><description><![CDATA[{popup_html}]]></description><styleUrl>{style_url}</styleUrl><Point><coordinates>{lon},{lat},0</coordinates></Point></Placemark>')
                    
                    # --- LINHA GUIA PARA A REDE ELÉTRICA MAIS PRÓXIMA ---
                    if pd.notna(rede_lat) and pd.notna(rede_lon):
                        rede_lat_str = str(rede_lat).replace(',', '.')
                        rede_lon_str = str(rede_lon).replace(',', '.')
                        kml_lines.append(f'        <Placemark><name>Guia de Rede: {dist_rede:.1f}m</name><styleUrl>#linha-ligacao-rede</styleUrl><LineString><tessellate>1</tessellate><coordinates>{lon},{lat},0 {rede_lon_str},{rede_lat_str},0</coordinates></LineString></Placemark>')

                kml_str_coords = "\n".join(coords_linha_kml)
                if kml_str_coords.strip():
                    kml_lines.append(f'        <Placemark><name>Contorno Rota</name><styleUrl>#linha-rota-contorno</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{kml_str_coords}\n            </coordinates></LineString></Placemark>')
                    kml_lines.append(f'        <Placemark><name>Traçado Rota</name><styleUrl>#rota-centro-{nome_limpo_base}</styleUrl><LineString><tessellate>1</tessellate><coordinates>\n{kml_str_coords}\n            </coordinates></LineString></Placemark>\n      </Folder>')
                else:
                    kml_lines.append('      </Folder>')
        kml_lines.append('  </Folder>')
    kml_lines.append('</Document>\n</kml>')
    return "\n".join(kml_lines)
